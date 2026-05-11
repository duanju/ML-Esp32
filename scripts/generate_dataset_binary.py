#!/usr/bin/env python3
"""Extract training dataset from Excel and generate binary + embedded C++ blob.

Reads all noise-level sheets, concatenates samples, and writes:
  - dataset.bin: raw float32 binary (inputs row-major, then targets)
  - dataset_blob.cpp: C++ const float array placed in flash RODATA

Input: Dataset_01(Noise Levels10%,20%,30%,40%,50%).xlsx
"""

import struct
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent / "main" / "data"
XLSX_PATH = BASE_DIR.parent.parent / "Dataset_01(Noise Levels10%,20%,30%,40%,50%).xlsx"
OUTPUT_BIN = BASE_DIR / "dataset.bin"
OUTPUT_BLOB_CPP = BASE_DIR / "dataset_blob.cpp"

NUM_FEATURES = 4


def format_float(v):
    """Format a Python float as a C++ float literal that preserves exact bits."""
    # Use %.9g for compact but precise representation
    s = "%.9g" % v
    # Ensure it has a decimal point or exponent for unambiguous float literal
    if "." not in s and "e" not in s and "E" not in s:
        s += ".0"
    return s + "f"


def main():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb["Sheet1"]

    inputs = []
    targets = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        f1, f2, f3, f4, label = row
        if f1 is None:
            break
        inputs.append([float(f1), float(f2), float(f3), float(f4)])
        targets.append(float(int(label)))

    num_samples = len(inputs)
    if num_samples == 0:
        print("ERROR: No data found in Sheet1")
        return 1

    print(
        "Extracted %d samples x %d features, %d targets"
        % (num_samples, NUM_FEATURES, len(targets))
    )

    all_floats = [v for sample in inputs for v in sample] + targets
    total_floats = len(all_floats)

    # --- Write binary ---
    raw = struct.pack("<%df" % total_floats, *all_floats)

    BASE_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_BIN, "wb") as f:
        f.write(raw)
    print("Written %d bytes to %s" % (len(raw), OUTPUT_BIN))

    # --- Generate C++ float array blob ---
    with open(OUTPUT_BLOB_CPP, "w") as f:
        f.write("// Auto-generated from Dataset_01.xlsx\n")
        f.write("// Do not edit by hand.\n")
        f.write('#include <cstddef>\n')
        f.write("namespace dataset {\n")
        f.write("extern const float blob[%d] = {\n" % total_floats)
        # Write 4 values per line (one sample + possibly overlapping targets)
        for i in range(0, total_floats, 4):
            chunk = all_floats[i : i + 4]
            line = ", ".join(format_float(v) for v in chunk)
            f.write("  " + line + ",\n")
        f.write("};\n")
        f.write("const size_t blob_size = %d;\n" % len(raw))
        f.write("const int num_samples = %d;\n" % num_samples)
        f.write("const int num_features = %d;\n" % NUM_FEATURES)
        f.write("}  // namespace dataset\n")
    print("Generated %s (%d floats)" % (OUTPUT_BLOB_CPP, total_floats))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
